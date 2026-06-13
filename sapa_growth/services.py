from __future__ import annotations

from datetime import date, timedelta
from io import BytesIO
from math import ceil
from typing import Any
from urllib.parse import quote, urlencode

from django.conf import settings
from django.http import Http404, HttpRequest, HttpResponse, HttpResponseBadRequest, HttpResponseNotAllowed
from openpyxl import Workbook
from reportlab.lib.utils import ImageReader
from reportlab.pdfgen import canvas as pdf_canvas

from etl.sapa_growth.control import log_export
from etl.sapa_growth.specs import GOLD_GLOBAL_SCHEMA, GOLD_SCHEMA
from etl.sapa_growth.storage import fetch_table, table_exists
from sapa_growth.logic import clean_text, map_course_status, parse_date
from sapa_growth.reporting import build_red_flag_rankings, build_video_rankings, compute_dashboard_metrics, course_status_counts, filter_rows
from sapa_growth.video_metadata import resolve_video_metadata, supported_video_link

SUMMARY_FIELDS = [
    "webinar_registrations_weekly",
    "webinar_registrations_cumulative",
    "webinar_registrations_previous",
    "field_rep_logins_weekly",
    "field_rep_logins_cumulative",
    "field_rep_logins_previous",
    "onboarded_doctors_weekly",
    "onboarded_doctors_cumulative",
    "onboarded_doctors_previous",
    "active_clinics_current",
    "active_clinics_cumulative",
    "active_clinics_previous",
    "inactive_clinics_current",
    "inactive_clinics_cumulative",
    "inactive_clinics_previous",
    "certified_clinics_current",
    "certified_clinics_cumulative",
    "certified_clinics_previous",
    "certified_clinics_supported",
    "total_screenings_weekly",
    "total_screenings_cumulative",
    "total_screenings_previous",
    "red_tags_weekly",
    "red_tags_cumulative",
    "red_tags_previous",
    "yellow_tags_weekly",
    "yellow_tags_cumulative",
    "yellow_tags_previous",
    "followups_scheduled_weekly",
    "followups_scheduled_cumulative",
    "followups_scheduled_previous",
    "reminders_sent_weekly",
    "reminders_sent_cumulative",
    "reminders_sent_previous",
    "doctor_course_started",
    "doctor_course_completed",
    "doctor_course_pending",
    "doctor_course_total",
    "paramedic_course_started",
    "paramedic_course_completed",
    "paramedic_course_pending",
    "paramedic_course_total",
]

DETAIL_SPECS = {
    "webinar_registrations": {
        "table": "rpt_webinar_registration_detail",
        "date_field": "registration_effective_date",
        "summary_date_field": "registration_effective_date",
        "title": "Webinar Registrations",
        "weekly": True,
        "columns": ["event_title", "start_date", "end_date", "timezone", "email", "first_name", "last_name", "doctor_display_name", "state", "field_rep_id"],
    },
    "onboarded_doctors": {
        "table": "rpt_doctor_status_current",
        "date_field": "first_seen_at",
        "summary_date_field": "first_seen_at",
        "summary_unique_field": "doctor_key",
        "title": "Onboarded Doctors",
        "weekly": True,
        "predicate": lambda row: row.get("onboarding_flag") == "true",
        "columns": [
            "campaign_label",
            "doctor_display_name",
            "city",
            "state",
            "field_rep_id",
            "doctor_has_logged_in",
            "doctor_has_updated_special_instructions",
            "doctor_has_added_clinic_staff",
            "clinic_staff_has_logged_in",
            "clinic_staff_forms_shared_count",
            "forms_filled_count",
            "red_tagged_patients_count",
            "yellow_tagged_patients_count",
            "registered_at",
        ],
    },
    "field_rep_logins": {
        "table": "rpt_field_rep_login_detail",
        "date_field": "login_ts",
        "summary_date_field": "login_ts",
        "summary_unique_field": "source_field_rep_id",
        "title": "Field Rep Logins",
        "weekly": True,
        "columns": ["field_rep_name", "field_rep_id", "state", "login_ts"],
    },
    "active_clinics": {
        "table": "rpt_doctor_status_current",
        "title": "Active Clinics",
        "weekly": False,
        "summary_mode": "status_history_active",
        "predicate": lambda row: row.get("active_flag") == "true",
        "columns": ["doctor_display_name", "city", "state", "field_rep_id", "screenings_last_15d", "last_screening_at"],
    },
    "inactive_clinics": {
        "table": "rpt_doctor_status_current",
        "title": "Inactive Clinics",
        "weekly": False,
        "summary_mode": "status_history_inactive",
        "predicate": lambda row: row.get("inactive_flag") == "true",
        "columns": ["doctor_display_name", "city", "state", "field_rep_id", "screenings_last_15d", "last_screening_at"],
    },
    "doctor_logins": {
        "table": "rpt_doctor_status_current",
        "title": "Doctor Logins",
        "weekly": False,
        "summary_mode": "current_doctor_login",
        "predicate": lambda row: clean_text(row.get("doctor_has_logged_in")).lower() == "yes",
        "columns": ["doctor_display_name", "city", "state", "field_rep_id", "doctor_has_logged_in"],
    },
    "certified_clinics": {
        "table": "rpt_certified_clinics",
        "title": "Certified Clinics",
        "weekly": False,
        "summary_mode": "status_history_certified",
        "columns": ["serial_order", "doctor_display_name", "city", "state", "field_rep_id", "certification_status", "certification_date"],
    },
    "total_screenings": {
        "table": "rpt_screening_detail",
        "date_field": "submitted_at",
        "summary_date_field": "submitted_at",
        "title": "Total Screenings",
        "weekly": True,
        "columns": ["doctor_display_name", "patient_id", "form_identifier", "language_code", "submitted_at", "overall_flag_code", "state", "field_rep_id"],
    },
    "red_tags": {
        "table": "rpt_tag_detail",
        "date_field": "submitted_at",
        "summary_date_field": "submitted_at",
        "title": "Red Tags",
        "weekly": True,
        "predicate": lambda row: (clean_text(row.get("tag_color")) or "").lower() == "red",
        "columns": ["doctor_display_name", "patient_id", "submitted_at", "tag_color", "individual_red_flag_count", "state", "field_rep_id"],
    },
    "yellow_tags": {
        "table": "rpt_tag_detail",
        "date_field": "submitted_at",
        "summary_date_field": "submitted_at",
        "title": "Yellow Tags",
        "weekly": True,
        "predicate": lambda row: (clean_text(row.get("tag_color")) or "").lower() == "yellow",
        "columns": ["doctor_display_name", "patient_id", "submitted_at", "tag_color", "state", "field_rep_id"],
    },
    "followups_scheduled": {
        "table": "rpt_followup_schedule_detail",
        "date_field": "scheduled_followup_date",
        "summary_date_field": "scheduled_followup_date",
        "title": "Follow-ups Scheduled",
        "weekly": True,
        "columns": ["doctor_display_name", "patient_id", "patient_name", "scheduled_followup_date", "schedule_sequence", "field_rep_id", "state"],
    },
    "reminders_sent": {
        "table": "rpt_reminder_sent_detail",
        "date_field": "ts",
        "summary_date_field": "ts",
        "title": "Reminders Sent",
        "weekly": True,
        "columns": ["doctor_display_name", "patient_id", "ts", "action_key", "field_rep_id", "state"],
    },
    "doctor_course_started": {
        "table": "rpt_course_detail",
        "title": "Doctor Course In Progress",
        "weekly": False,
        "summary_date_field": "started_at",
        "predicate": lambda row: clean_text(row.get("course_audience")) == "doctor" and map_course_status(row.get("dashboard_status") or row.get("progress_status")) == "In Progress",
        "columns": ["display_name", "user_email", "phone", "progress_status", "enrolled_at", "started_at", "doctor_display_name", "state", "field_rep_id"],
    },
    "doctor_course_completed": {
        "table": "rpt_course_detail",
        "title": "Doctor Course Completed",
        "weekly": False,
        "summary_date_field": "completed_at",
        "predicate": lambda row: clean_text(row.get("course_audience")) == "doctor" and map_course_status(row.get("dashboard_status") or row.get("progress_status")) == "Completed",
        "columns": ["display_name", "user_email", "phone", "progress_status", "completed_at", "doctor_display_name", "state", "field_rep_id"],
    },
    "doctor_course_pending": {
        "table": "rpt_course_detail",
        "title": "Doctor Course Not Started",
        "weekly": False,
        "summary_date_field": "enrolled_at",
        "predicate": lambda row: clean_text(row.get("course_audience")) == "doctor" and map_course_status(row.get("dashboard_status") or row.get("progress_status")) == "Not Started",
        "columns": ["display_name", "user_email", "phone", "progress_status", "enrolled_at", "doctor_display_name", "state", "field_rep_id"],
    },
    "paramedic_course_started": {
        "table": "rpt_course_detail",
        "title": "Paramedic Course In Progress",
        "weekly": False,
        "summary_date_field": "started_at",
        "predicate": lambda row: clean_text(row.get("course_audience")) == "paramedic" and map_course_status(row.get("dashboard_status") or row.get("progress_status")) == "In Progress",
        "columns": ["display_name", "user_email", "phone", "progress_status", "enrolled_at", "started_at", "doctor_display_name", "state", "field_rep_id"],
    },
    "paramedic_course_completed": {
        "table": "rpt_course_detail",
        "title": "Paramedic Course Completed",
        "weekly": False,
        "summary_date_field": "completed_at",
        "predicate": lambda row: clean_text(row.get("course_audience")) == "paramedic" and map_course_status(row.get("dashboard_status") or row.get("progress_status")) == "Completed",
        "columns": ["display_name", "user_email", "phone", "progress_status", "completed_at", "doctor_display_name", "state", "field_rep_id"],
    },
    "paramedic_course_pending": {
        "table": "rpt_course_detail",
        "title": "Paramedic Course Not Started",
        "weekly": False,
        "summary_date_field": "enrolled_at",
        "predicate": lambda row: clean_text(row.get("course_audience")) == "paramedic" and map_course_status(row.get("dashboard_status") or row.get("progress_status")) == "Not Started",
        "columns": ["display_name", "user_email", "phone", "progress_status", "enrolled_at", "doctor_display_name", "state", "field_rep_id"],
    },
    "patient_videos": {
        "table": "rpt_video_view_detail",
        "title": "Top Patient Education Videos Viewed",
        "weekly": False,
        "summary_date_field": "ts",
        "predicate": lambda row: clean_text(row.get("audience")) == "patient",
        "columns": ["rank", "preferred_display_label", "view_count", "latest_interaction_timestamp"],
    },
    "doctor_videos": {
        "table": "rpt_video_view_detail",
        "title": "Top Doctor Education Videos Viewed",
        "weekly": False,
        "summary_date_field": "ts",
        "predicate": lambda row: clean_text(row.get("audience")) == "doctor",
        "columns": ["rank", "preferred_display_label", "view_count", "latest_interaction_timestamp"],
    },
}

DETAIL_WINDOWS = [
    {"key": "last_24_hours", "label": "Last 24 Hours", "days": 0},
    {"key": "last_week", "label": "Last Week", "days": 6},
    {"key": "last_month", "label": "Last Month", "days": 29},
    {"key": "cumulative", "label": "Cumulative", "days": None},
]

DETAIL_WINDOW_KEYS = {str(window["key"]) for window in DETAIL_WINDOWS}


def _scope_campaign_key(scope: Any = None) -> str:
    if isinstance(scope, dict):
        return clean_text(scope.get("campaign_key"))
    return clean_text(scope)


def _global_rows(table: str) -> list[dict[str, Any]]:
    if not table_exists(GOLD_GLOBAL_SCHEMA, table):
        return []
    return fetch_table(GOLD_GLOBAL_SCHEMA, table)


def _campaign_registry_row(campaign_key: Any) -> dict[str, Any] | None:
    key = clean_text(campaign_key)
    if not key:
        return None
    normalized = "".join(ch.lower() for ch in key if ch.isalnum())
    for row in _global_rows("campaign_registry"):
        if clean_text(row.get("campaign_key")) == key:
            return row
        if clean_text(row.get("campaign_id_normalized")) == normalized:
            return row
    return None


def _gold_rows(table: str, scope: Any = None) -> list[dict[str, Any]]:
    campaign_key = _scope_campaign_key(scope)
    if campaign_key:
        registry = _campaign_registry_row(campaign_key)
        schema = clean_text((registry or {}).get("gold_schema_name"))
        if not schema or not table_exists(schema, table):
            return []
        return fetch_table(schema, table)
    if table_exists(GOLD_SCHEMA, table):
        return fetch_table(GOLD_SCHEMA, table)
    return []


def _supported_video_link(value: Any) -> str:
    return supported_video_link(value)


def _resolved_video_link(row: dict[str, Any]) -> str:
    return (
        _supported_video_link(row.get("video_url"))
        or _supported_video_link(row.get("content_identifier"))
        or _supported_video_link(row.get("preferred_display_label"))
        or ""
    )


def _enrich_red_flag_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    enriched = []
    for row in rows:
        item = dict(row)
        item["red_flag_name"] = clean_text(row.get("red_flag_name")) or clean_text(row.get("red_flag")) or ""
        enriched.append(item)
    return enriched


def _enrich_video_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    enriched = []
    for row in rows:
        item = dict(row)
        resolved_link = _resolved_video_link(item)
        if not resolved_link:
            continue
        metadata = resolve_video_metadata(resolved_link)
        video_title = metadata["video_title"] or clean_text(row.get("video_title")) or clean_text(row.get("preferred_display_label"))
        item["content_identifier"] = resolved_link
        item["video_url"] = resolved_link
        item["video_title"] = video_title if video_title and video_title != resolved_link else ""
        item["preferred_display_label"] = metadata["preferred_display_label"] or item["video_title"] or resolved_link
        enriched.append(item)
    return enriched


def _default_campaign() -> dict[str, str]:
    return {
        "underlying_key": clean_text(settings.SAPA_DASHBOARD.get("DEFAULT_CAMPAIGN_KEY")) or "growth-clinic",
        "display_label": clean_text(settings.SAPA_DASHBOARD.get("DEFAULT_CAMPAIGN_LABEL")) or "SAPA Growth Clinic Program",
    }


def _campaign_route_base(filters: dict[str, str | None]) -> str:
    campaign_key = clean_text(filters.get("campaign_key"))
    if campaign_key:
        return f"/sapa-growth/campaign/{quote(campaign_key)}/"
    return "/sapa-growth/"


def _metric_href(metric: str, filters: dict[str, str | None], extra_query: dict[str, str] | None = None) -> str:
    base = _campaign_route_base(filters)
    query_parts = [current_filters_query(filters, include_campaign=not clean_text(filters.get("campaign_key")))]
    if extra_query:
        query_parts.append(urlencode({key: value for key, value in extra_query.items() if value}))
    query_string = "&".join(part for part in query_parts if part)
    return f"{base}details/{metric}/" + (f"?{query_string}" if query_string else "")


def _normalize_detail_window(value: Any) -> str:
    window_key = clean_text(value)
    return window_key if window_key in DETAIL_WINDOW_KEYS else ""


def _detail_window_href(metric: str, filters: dict[str, str | None], window_key: str) -> str:
    base = f"{_campaign_route_base(filters)}details/{metric}/"
    filters_query = current_filters_query(filters, include_campaign=not clean_text(filters.get("campaign_key")))
    window_query = urlencode({"window": window_key})
    return f"{base}?{filters_query}&{window_query}" if filters_query else f"{base}?{window_query}"


def _latest_refresh(scope: Any = None) -> dict[str, Any] | None:
    rows = _gold_rows("refresh_status", scope) or _global_rows("refresh_status")
    if not rows:
        return None
    rows.sort(key=lambda row: row.get("published_at") or "", reverse=True)
    return rows[0]


def _to_int(value: Any) -> int:
    try:
        return int(float(str(value)))
    except (TypeError, ValueError):
        return 0


def _summary_from_row(row: dict[str, Any]) -> dict[str, Any]:
    summary = {}
    for field in SUMMARY_FIELDS:
        if field == "certified_clinics_supported":
            summary[field] = "true" if clean_text(row.get(field)) == "true" else "false"
        else:
            summary[field] = _to_int(row.get(field))
    summary["as_of_date"] = clean_text(row.get("as_of_date")) or date.today().isoformat()
    summary["weekly_window_start"] = clean_text(row.get("weekly_window_start")) or summary["as_of_date"]
    summary["weekly_window_end"] = clean_text(row.get("weekly_window_end")) or summary["as_of_date"]
    summary["activity_window_start"] = clean_text(row.get("activity_window_start")) or summary["as_of_date"]
    summary["activity_window_end"] = clean_text(row.get("activity_window_end")) or summary["as_of_date"]
    summary["published_at"] = clean_text(row.get("published_at")) or ""
    return summary


def _metric_ready_doctor_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    output = []
    for row in rows:
        item = dict(row)
        item["is_user_created_doctor"] = clean_text(row.get("is_user_created_doctor")) or clean_text(row.get("onboarding_flag")) or "false"
        output.append(item)
    return output


def _metric_ready_status_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    output = []
    for row in rows:
        item = dict(row)
        item["is_active"] = clean_text(row.get("is_active")) or clean_text(row.get("active_flag")) or "false"
        item["is_inactive"] = clean_text(row.get("is_inactive")) or clean_text(row.get("inactive_flag")) or "false"
        output.append(item)
    return output


def _doctor_course_enrollments(course_rows: list[dict[str, Any]]) -> dict[str, str]:
    enrollments: dict[str, str] = {}
    for row in course_rows:
        if clean_text(row.get("course_audience")) != "doctor":
            continue
        doctor_key = clean_text(row.get("doctor_key"))
        if not doctor_key:
            continue
        enrolled_at = clean_text(row.get("enrolled_at")) or ""
        if doctor_key not in enrollments or (enrolled_at and enrolled_at < enrollments[doctor_key]):
            enrollments[doctor_key] = enrolled_at
    return enrollments


def _derived_certified_summary(
    filters: dict[str, str | None],
    refresh: dict[str, Any] | None,
    doctor_rows: list[dict[str, Any]],
    doctor_history_rows: list[dict[str, Any]],
    course_rows: list[dict[str, Any]],
) -> dict[str, int | str]:
    enrolled = _doctor_course_enrollments(filter_rows(course_rows, filters))
    current_rows = filter_rows(doctor_rows, filters)
    current = len(
        {
            clean_text(row.get("doctor_key"))
            for row in current_rows
            if row.get("active_flag") == "true" and clean_text(row.get("doctor_key")) in enrolled
        }
        - {None}
    )
    as_of = parse_date((refresh or {}).get("as_of_date")) or date.today()
    previous_date = (as_of - timedelta(days=1)).isoformat()
    history_rows = filter_rows(doctor_history_rows, filters)
    previous = len(
        {
            clean_text(row.get("doctor_key"))
            for row in history_rows
            if row.get("as_of_date") == previous_date and row.get("is_active") == "true" and clean_text(row.get("doctor_key")) in enrolled
        }
        - {None}
    )
    cumulative = len(
        {
            clean_text(row.get("doctor_key"))
            for row in history_rows
            if row.get("is_active") == "true" and clean_text(row.get("doctor_key")) in enrolled
        }
        - {None}
    )
    return {
        "certified_clinics_current": current,
        "certified_clinics_previous": previous,
        "certified_clinics_cumulative": cumulative,
        "certified_clinics_supported": "true",
    }


def _derived_certified_rows(
    filters: dict[str, str | None],
    doctor_rows: list[dict[str, Any]],
    course_rows: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    enrolled = _doctor_course_enrollments(filter_rows(course_rows, filters))
    filtered_doctors = filter_rows(doctor_rows, filters)
    rows: list[dict[str, Any]] = []
    for doctor in filtered_doctors:
        doctor_key = clean_text(doctor.get("doctor_key"))
        if doctor.get("active_flag") != "true" or not doctor_key or doctor_key not in enrolled:
            continue
        rows.append(
            {
                "serial_order": str(len(rows) + 1),
                "doctor_key": doctor_key,
                "campaign_key": doctor.get("campaign_key", ""),
                "campaign_label": doctor.get("campaign_label", ""),
                "doctor_display_name": doctor.get("doctor_display_name", ""),
                "city": doctor.get("city", ""),
                "district": doctor.get("district", ""),
                "state": doctor.get("state", ""),
                "field_rep_id": doctor.get("field_rep_id", ""),
                "field_rep_name": doctor.get("field_rep_name", ""),
                "certification_status": "enrolled",
                "certification_date": enrolled.get(doctor_key, ""),
                "certification_source": "doctor_course_enrollment",
            }
        )
    return rows


def parse_global_filters(query_params: Any, campaign_key: str | None = None) -> dict[str, str | None]:
    return {
        "campaign_key": clean_text(campaign_key) or clean_text(query_params.get("campaign_key")),
        "state": clean_text(query_params.get("state")),
        "field_rep_id": clean_text(query_params.get("field_rep_id")),
        "doctor_key": clean_text(query_params.get("doctor_key")),
    }


def parse_certified_filters(query_params: Any, global_filters: dict[str, str | None]) -> dict[str, str | None]:
    return {
        "campaign_key": global_filters.get("campaign_key"),
        "state": global_filters.get("state"),
        "field_rep_id": clean_text(query_params.get("cert_field_rep_id")) or global_filters.get("field_rep_id"),
        "doctor_key": clean_text(query_params.get("cert_doctor_key")) or global_filters.get("doctor_key"),
        "city": clean_text(query_params.get("cert_city")),
    }


def current_filters_query(filters: dict[str, str | None], include_campaign: bool = True) -> str:
    payload = {key: value for key, value in filters.items() if value}
    if not include_campaign:
        payload.pop("campaign_key", None)
    return urlencode(payload)


def _with_delta(current: int | None, previous: int | None) -> dict[str, Any]:
    if current is None or previous is None:
        return {"value": current, "delta": None}
    return {"value": current, "delta": current - previous}


def _dashboard_tiles(summary: dict[str, Any], filters: dict[str, str | None]) -> dict[str, list[dict[str, Any]]]:
    return {
        "field_rep": [
            # Webinar registrations are temporarily hidden until campaign attribution is finalized.
            {
                "title": "Onboarded Doctors",
                "value": summary["onboarded_doctors_weekly"],
                "cumulative": summary["onboarded_doctors_cumulative"],
                "delta": summary["onboarded_doctors_weekly"] - summary["onboarded_doctors_previous"],
                "href": _metric_href("onboarded_doctors", filters),
                "theme": "teal",
                "supported": True,
            },
            {
                "title": "Field Rep Logins",
                "value": summary.get("field_rep_logins_weekly", 0),
                "cumulative": summary.get("field_rep_logins_cumulative", 0),
                "delta": summary.get("field_rep_logins_weekly", 0) - summary.get("field_rep_logins_previous", 0),
                "href": _metric_href("field_rep_logins", filters),
                "theme": "teal",
                "supported": True,
            },
        ],
        "clinic": [
            {
                "title": "Active Clinics",
                "value": summary["active_clinics_current"],
                "cumulative": summary["active_clinics_cumulative"],
                "delta": summary["active_clinics_current"] - summary["active_clinics_previous"],
                "href": _metric_href("active_clinics", filters),
                "theme": "positive",
                "supported": True,
            },
            {
                "title": "Inactive Clinics",
                "value": summary["inactive_clinics_current"],
                "cumulative": summary["inactive_clinics_cumulative"],
                "delta": summary["inactive_clinics_current"] - summary["inactive_clinics_previous"],
                "href": _metric_href("inactive_clinics", filters),
                "theme": "negative",
                "supported": True,
            },
            {
                "title": "Certified Clinics",
                "value": summary["certified_clinics_current"],
                "cumulative": summary["certified_clinics_cumulative"],
                "delta": 0 if summary["certified_clinics_current"] is None or summary["certified_clinics_previous"] is None else summary["certified_clinics_current"] - summary["certified_clinics_previous"],
                "href": _metric_href("certified_clinics", filters),
                "theme": "neutral",
                "supported": True,
            },
            {
                "title": "Doctor Logins",
                "value": summary.get("doctor_logins_current", 0),
                "cumulative": summary.get("doctor_logins_current", 0),
                "delta": 0,
                "href": _metric_href("doctor_logins", filters, {"window": "cumulative"}),
                "theme": "teal",
                "supported": True,
            },
            {
                "title": "Total Screenings",
                "value": summary["total_screenings_weekly"],
                "cumulative": summary["total_screenings_cumulative"],
                "delta": summary["total_screenings_weekly"] - summary["total_screenings_previous"],
                "href": _metric_href("total_screenings", filters),
                "theme": "neutral",
                "supported": True,
            },
            {
                "title": "Red Tags",
                "value": summary["red_tags_weekly"],
                "cumulative": summary["red_tags_cumulative"],
                "delta": summary["red_tags_weekly"] - summary["red_tags_previous"],
                "href": _metric_href("red_tags", filters),
                "theme": "warning",
                "supported": True,
            },
            {
                "title": "Yellow Tags",
                "value": summary["yellow_tags_weekly"],
                "cumulative": summary["yellow_tags_cumulative"],
                "delta": summary["yellow_tags_weekly"] - summary["yellow_tags_previous"],
                "href": _metric_href("yellow_tags", filters),
                "theme": "warning",
                "supported": True,
            },
        ],
        "operational": [
            {
                "title": "Follow-ups Scheduled",
                "value": summary["followups_scheduled_weekly"],
                "cumulative": summary["followups_scheduled_cumulative"],
                "delta": summary["followups_scheduled_weekly"] - summary["followups_scheduled_previous"],
                "href": _metric_href("followups_scheduled", filters),
                "theme": "teal",
                "supported": True,
            },
            {
                "title": "Reminders Sent",
                "value": summary["reminders_sent_weekly"],
                "cumulative": summary["reminders_sent_cumulative"],
                "delta": summary["reminders_sent_weekly"] - summary["reminders_sent_previous"],
                "href": _metric_href("reminders_sent", filters),
                "theme": "teal",
                "supported": True,
            },
        ],
    }


def _course_cards(course_rows: list[dict[str, Any]], filters: dict[str, str | None]) -> list[dict[str, Any]]:
    counts = course_status_counts(course_rows)
    cards = []
    for audience, title in (("doctor", "Doctor Course"), ("paramedic", "Paramedic Course")):
        summary = counts.get(audience, {"Not Started": 0, "In Progress": 0, "Completed": 0, "Total": 0})
        total = summary["Total"] or 0
        rows = []
        for status, metric_suffix in (("Not Started", "pending"), ("In Progress", "started"), ("Completed", "completed")):
            metric_key = f"{audience}_course_{metric_suffix}"
            rows.append(
                {
                    "label": status,
                    "count": summary[status],
                    "ratio": round((summary[status] / total) * 100, 2) if total else 0,
                    "href": _metric_href(metric_key, filters, {"window": "cumulative"}),
                }
            )
        cards.append({"title": title, "rows": rows, "total": total})
    return cards


def _doctor_login_count(doctor_rows: list[dict[str, Any]]) -> int:
    return len(
        {
            clean_text(row.get("doctor_key"))
            for row in doctor_rows
            if clean_text(row.get("doctor_key")) and clean_text(row.get("doctor_has_logged_in")).lower() == "yes"
        }
    )


def _state_label(row: dict[str, Any]) -> str:
    return clean_text(row.get("state")) or "Unknown"


def _state_performance_rows(
    doctor_rows: list[dict[str, Any]],
    screening_rows: list[dict[str, Any]],
    field_rep_login_rows: list[dict[str, Any]],
    filters: dict[str, str | None],
) -> list[dict[str, Any]]:
    filtered_doctors = filter_rows(doctor_rows, filters)
    filtered_screenings = filter_rows(screening_rows, filters)
    filtered_field_rep_logins = filter_rows(field_rep_login_rows, filters)
    onboarded_by_state: dict[str, set[str]] = {}
    screenings_by_state: dict[str, int] = {}
    field_rep_logins_by_state: dict[str, set[str]] = {}
    for row in filtered_doctors:
        if clean_text(row.get("onboarding_flag")) != "true":
            continue
        state = _state_label(row)
        doctor_key = clean_text(row.get("doctor_key"))
        if doctor_key:
            onboarded_by_state.setdefault(state, set()).add(doctor_key)
    for row in filtered_screenings:
        state = _state_label(row)
        screenings_by_state[state] = screenings_by_state.get(state, 0) + 1
    for row in filtered_field_rep_logins:
        state = _state_label(row)
        field_rep_key = clean_text(row.get("source_field_rep_id")) or clean_text(row.get("field_rep_id"))
        if field_rep_key:
            field_rep_logins_by_state.setdefault(state, set()).add(field_rep_key)
    states = sorted(set(onboarded_by_state) | set(screenings_by_state) | set(field_rep_logins_by_state), key=lambda value: (value == "Unknown", value.lower()))
    max_onboarded = max((len(onboarded_by_state.get(state, set())) for state in states), default=0) or 1
    max_screenings = max((screenings_by_state.get(state, 0) for state in states), default=0) or 1
    max_field_rep_logins = max((len(field_rep_logins_by_state.get(state, set())) for state in states), default=0) or 1
    return [
        {
            "state": state,
            "onboarded_doctors": len(onboarded_by_state.get(state, set())),
            "screenings": screenings_by_state.get(state, 0),
            "field_rep_logins": len(field_rep_logins_by_state.get(state, set())),
            "onboarded_pct": round((len(onboarded_by_state.get(state, set())) / max_onboarded) * 100, 2),
            "screenings_pct": round((screenings_by_state.get(state, 0) / max_screenings) * 100, 2),
            "field_rep_logins_pct": round((len(field_rep_logins_by_state.get(state, set())) / max_field_rep_logins) * 100, 2),
        }
        for state in states
    ]


def filter_options(scope: Any = None) -> dict[str, list[dict[str, Any]]]:
    campaign_key = _scope_campaign_key(scope)
    return {
        "campaigns": campaign_options(),
        "states": _gold_rows("dim_filter_state", campaign_key),
        "field_reps": _gold_rows("dim_filter_field_rep", campaign_key),
        "doctors": _gold_rows("dim_filter_doctor", campaign_key),
        "cities": _gold_rows("dim_filter_city", campaign_key),
    }


def _sorted_campaign_options(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    unique: dict[str, dict[str, Any]] = {}
    for row in rows:
        key = clean_text(row.get("underlying_key") or row.get("campaign_key"))
        label = clean_text(row.get("display_label") or row.get("campaign_label"))
        if not key:
            continue
        unique[key] = {
            "underlying_key": key,
            "display_label": label or key,
        }
    return sorted(unique.values(), key=lambda item: ((item.get("display_label") or "").lower(), (item.get("underlying_key") or "").lower()))


def campaign_options() -> list[dict[str, Any]]:
    registry_campaigns = _sorted_campaign_options(
        [
            {
                "underlying_key": row.get("campaign_key"),
                "display_label": row.get("campaign_label"),
            }
            for row in _global_rows("campaign_registry")
        ]
    )
    if registry_campaigns:
        return registry_campaigns
    campaigns = _sorted_campaign_options(_gold_rows("dim_filter_campaign"))
    if campaigns:
        return campaigns
    campaigns = _sorted_campaign_options(_gold_rows("dashboard_summary_state_rep"))
    if campaigns:
        return campaigns
    campaigns = _sorted_campaign_options(_gold_rows("rpt_doctor_status_current"))
    if campaigns:
        return campaigns
    if table_exists("raw_sapa_mysql", "campaign_campaign_raw"):
        master_campaigns = _sorted_campaign_options(
            [
                {
                    "underlying_key": row.get("id"),
                    "display_label": row.get("name"),
                }
                for row in fetch_table("raw_sapa_mysql", "campaign_campaign_raw")
                if clean_text(row.get("system_rfa")).lower() in {"1", "true", "t", "yes"}
                and clean_text(row.get("status")).lower() == "published"
                and clean_text(row.get("id"))
                and clean_text(row.get("name"))
                and "test" not in clean_text(row.get("name")).lower()
                and "dummy" not in clean_text(row.get("name")).lower()
            ]
        )
        if master_campaigns:
            return master_campaigns
    return [_default_campaign()]


def dashboard_context(filters: dict[str, str | None]) -> dict[str, Any]:
    campaign_key = clean_text(filters.get("campaign_key"))
    refresh = _latest_refresh(campaign_key)
    options = filter_options(campaign_key)
    selected_campaign = next(
        (option for option in options["campaigns"] if clean_text(option.get("underlying_key")) == campaign_key),
        None,
    )
    if refresh is None:
        return {
            "ready": False,
            "filters": filters,
            "filter_options": options,
            "export_filename": "sapa-growth-dashboard-report.pdf",
            "campaign": selected_campaign or _default_campaign(),
            "route_base": _campaign_route_base(filters),
        }

    summary = None
    snapshot_rows = _gold_rows("dashboard_summary_snapshot", campaign_key)
    if not any(filters.values()) and snapshot_rows:
        summary = _summary_from_row(snapshot_rows[0])
    elif campaign_key and not any(value for key, value in filters.items() if key != "campaign_key") and snapshot_rows:
        summary = _summary_from_row(snapshot_rows[0])
    elif not filters.get("doctor_key"):
        helper_rows = _gold_rows("dashboard_summary_state_rep", campaign_key)
        for row in helper_rows:
            if (
                clean_text(row.get("campaign_key")) == campaign_key
                and clean_text(row.get("state")) == clean_text(filters.get("state"))
                and clean_text(row.get("field_rep_id")) == clean_text(filters.get("field_rep_id"))
            ):
                summary = _summary_from_row(row)
                break

    doctor_rows = _gold_rows("rpt_doctor_status_current", campaign_key)
    doctor_history_rows = _gold_rows("rpt_doctor_status_history", campaign_key)
    screening_rows = _gold_rows("rpt_screening_detail", campaign_key)
    raw_redflag_rows = _gold_rows("rpt_submission_redflag_detail", campaign_key)
    followup_rows = _gold_rows("rpt_followup_schedule_detail", campaign_key)
    reminder_rows = _gold_rows("rpt_reminder_sent_detail", campaign_key)
    field_rep_login_rows = _gold_rows("rpt_field_rep_login_detail", campaign_key)
    webinar_rows = _gold_rows("rpt_webinar_registration_detail", campaign_key)
    course_rows = _gold_rows("rpt_course_detail", campaign_key)
    certified_rows = _derived_certified_rows(filters, doctor_rows, course_rows)

    if summary is None:
        as_of = parse_date(refresh.get("as_of_date")) or date.today()
        metric_doctor_rows = _metric_ready_doctor_rows(doctor_rows)
        metric_status_rows = _metric_ready_status_rows(doctor_rows)
        summary = compute_dashboard_metrics(
            as_of_date=as_of,
            doctor_rows=filter_rows(metric_doctor_rows, filters),
            doctor_status_current_rows=filter_rows(metric_status_rows, filters),
            doctor_status_history_rows=filter_rows(doctor_history_rows, filters),
            certification_rows=[],
            webinar_rows=filter_rows(webinar_rows, filters),
            screening_rows=filter_rows(screening_rows, filters),
            followup_rows=filter_rows(followup_rows, filters),
            reminder_rows=filter_rows(reminder_rows, filters),
            course_rows=filter_rows(course_rows, filters),
            field_rep_login_rows=filter_rows(field_rep_login_rows, filters),
        )
        summary["published_at"] = refresh.get("published_at") or ""

    summary.update(_derived_certified_summary(filters, refresh, doctor_rows, doctor_history_rows, course_rows))

    filtered_course_rows = filter_rows(course_rows, filters)
    filtered_redflag_rows = _enrich_red_flag_rows(filter_rows(raw_redflag_rows, filters))
    filtered_doctor_rows = filter_rows(doctor_rows, filters)
    summary["doctor_logins_current"] = _doctor_login_count(filtered_doctor_rows)
    filtered_field_rep_login_rows = filter_rows(field_rep_login_rows, filters)
    as_of_for_logins = parse_date(refresh.get("as_of_date")) or date.today()
    field_rep_login_counts = _count_window(
        filtered_field_rep_login_rows,
        date_field="login_ts",
        as_of=as_of_for_logins,
        unique_field="source_field_rep_id",
    )
    summary["field_rep_logins_weekly"] = field_rep_login_counts["Last Week"]
    summary["field_rep_logins_cumulative"] = field_rep_login_counts["Cumulative"]
    summary["field_rep_logins_previous"] = max(summary.get("field_rep_logins_previous", 0), 0)
    certified_supported = True
    filters_query = current_filters_query(filters, include_campaign=not clean_text(filters.get("campaign_key")))
    campaign = selected_campaign or _default_campaign()

    return {
        "ready": True,
        "refresh": refresh,
        "summary": summary,
        "export_filename": f"sapa-growth-dashboard-{summary.get('as_of_date') or refresh.get('as_of_date') or 'report'}.pdf",
        "campaign": campaign,
        "route_base": _campaign_route_base(filters),
        "dashboard_export_href": f"{_campaign_route_base(filters)}export/dashboard.pdf" + (f"?{filters_query}" if filters_query else ""),
        "tiles": _dashboard_tiles(summary, filters),
        "course_cards": _course_cards(filtered_course_rows, filters),
        "state_performance": _state_performance_rows(doctor_rows, screening_rows, field_rep_login_rows, filters),
        "red_flag_rankings": build_red_flag_rankings(filtered_redflag_rows),
        "filters": filters,
        "filters_query": filters_query,
        "filter_options": options,
        "certified_supported": certified_supported,
        "certified_toggle_url": f"{_campaign_route_base(filters)}certified-clinics/" + (f"?{filters_query}" if filters_query else ""),
        "certified_rows": certified_rows,
    }


def certified_context(global_filters: dict[str, str | None], local_filters: dict[str, str | None]) -> dict[str, Any]:
    campaign_key = clean_text(local_filters.get("campaign_key") or global_filters.get("campaign_key"))
    rows = _derived_certified_rows(
        local_filters,
        _gold_rows("rpt_doctor_status_current", campaign_key),
        _gold_rows("rpt_course_detail", campaign_key),
    )
    return {
        "supported": True,
        "rows": rows,
        "filters": local_filters,
        "filter_options": filter_options(campaign_key),
        "route_base": _campaign_route_base(global_filters),
        "export_query": current_filters_query(
            {
                "campaign_key": global_filters.get("campaign_key"),
                "state": global_filters.get("state"),
                "cert_field_rep_id": local_filters.get("field_rep_id"),
                "cert_doctor_key": local_filters.get("doctor_key"),
                "cert_city": local_filters.get("city"),
            }
        ),
    }


def _count_window(
    rows: list[dict[str, Any]],
    *,
    date_field: str,
    as_of: date,
    unique_field: str | None = None,
    predicate: Any = None,
    include_undated_cumulative: bool = False,
) -> dict[str, int]:
    def _window(start: date | None = None, end: date | None = None) -> int:
        bounded_window = start is not None or end is not None
        if unique_field:
            keys: set[str] = set()
            for row in rows:
                if predicate and not predicate(row):
                    continue
                row_date = parse_date(row.get(date_field))
                if row_date is None:
                    if bounded_window or not include_undated_cumulative:
                        continue
                if start and row_date < start:
                    continue
                if end and row_date > end:
                    continue
                key = clean_text(row.get(unique_field))
                if key:
                    keys.add(key)
            return len(keys)
        total = 0
        for row in rows:
            if predicate and not predicate(row):
                continue
            row_date = parse_date(row.get(date_field))
            if row_date is None:
                if bounded_window or not include_undated_cumulative:
                    continue
            if start and row_date < start:
                continue
            if end and row_date > end:
                continue
            total += 1
        return total

    return {
        "Last 24 Hours": _window(as_of, as_of),
        "Last Week": _window(as_of - timedelta(days=6), as_of),
        "Last Month": _window(as_of - timedelta(days=29), as_of),
        "Cumulative": _window(),
    }


def _window_bounds(window_key: str, as_of: date) -> tuple[date | None, date | None]:
    normalized = _normalize_detail_window(window_key)
    if normalized == "cumulative":
        return None, None
    days = next((window["days"] for window in DETAIL_WINDOWS if window["key"] == normalized), None)
    if days is None:
        return None, None
    return as_of - timedelta(days=int(days)), as_of


def _filter_rows_by_detail_window(
    rows: list[dict[str, Any]],
    date_field: str | None,
    as_of: date,
    window_key: str,
    *,
    include_undated_cumulative: bool = False,
) -> list[dict[str, Any]]:
    if not date_field:
        return rows
    start, end = _window_bounds(window_key, as_of)
    bounded_window = start is not None or end is not None
    filtered = []
    for row in rows:
        row_date = parse_date(row.get(date_field))
        if row_date is None:
            if bounded_window or not include_undated_cumulative:
                continue
        if start and row_date < start:
            continue
        if end and row_date > end:
            continue
        filtered.append(row)
    return filtered


def _dedupe_latest_rows(rows: list[dict[str, Any]], key_field: str, date_field: str) -> list[dict[str, Any]]:
    latest: dict[str, dict[str, Any]] = {}
    for row in sorted(rows, key=lambda item: item.get(date_field) or "", reverse=True):
        key = clean_text(row.get(key_field))
        if key and key not in latest:
            latest[key] = row
    return list(latest.values())


def _status_history_rows_for_window(
    filters: dict[str, str | None],
    *,
    predicate: Any,
    as_of: date,
    window_key: str,
) -> list[dict[str, Any]]:
    rows = filter_rows(_gold_rows("rpt_doctor_status_history", filters), filters)
    rows = _filter_rows_by_detail_window(rows, "as_of_date", as_of, window_key)
    rows = [row for row in rows if predicate(row)]
    output = []
    for row in _dedupe_latest_rows(rows, "doctor_key", "as_of_date"):
        item = dict(row)
        item["active_flag"] = clean_text(row.get("is_active"))
        item["inactive_flag"] = clean_text(row.get("is_inactive"))
        output.append(item)
    return output


def _certified_rows_for_window(filters: dict[str, str | None], as_of: date, window_key: str) -> list[dict[str, Any]]:
    history_rows = filter_rows(_gold_rows("rpt_doctor_status_history", filters), filters)
    history_rows = _filter_rows_by_detail_window(history_rows, "as_of_date", as_of, window_key)
    course_rows = filter_rows(_gold_rows("rpt_course_detail", filters), filters)
    enrolled = set(_doctor_course_enrollments(course_rows).keys())
    rows = [
        row
        for row in history_rows
        if row.get("is_active") == "true" and clean_text(row.get("doctor_key")) in enrolled
    ]
    output = []
    for index, row in enumerate(_dedupe_latest_rows(rows, "doctor_key", "as_of_date"), start=1):
        item = dict(row)
        item["serial_order"] = str(index)
        item["certification_status"] = "Course enrolled and active"
        item["certification_date"] = clean_text(row.get("as_of_date"))
        item["certification_source"] = "Doctor course enrollment"
        output.append(item)
    return output


def _status_history_window_counts(
    filters: dict[str, str | None],
    *,
    predicate: Any,
    as_of: date,
) -> dict[str, int]:
    rows = filter_rows(_gold_rows("rpt_doctor_status_history", filters), filters)
    return _count_window(rows, date_field="as_of_date", as_of=as_of, unique_field="doctor_key", predicate=predicate)


def _certified_window_counts(filters: dict[str, str | None], as_of: date) -> dict[str, int]:
    history_rows = filter_rows(_gold_rows("rpt_doctor_status_history", filters), filters)
    course_rows = filter_rows(_gold_rows("rpt_course_detail", filters), filters)
    enrolled = set(_doctor_course_enrollments(course_rows).keys())
    return _count_window(
        history_rows,
        date_field="as_of_date",
        as_of=as_of,
        unique_field="doctor_key",
        predicate=lambda row: row.get("is_active") == "true" and clean_text(row.get("doctor_key")) in enrolled,
    )


def _metric_summary_cards(metric: str, filters: dict[str, str | None], refresh: dict[str, Any], selected_window: str = "") -> list[dict[str, Any]]:
    spec = DETAIL_SPECS[metric]
    as_of = parse_date(refresh.get("as_of_date")) or date.today()
    summary_mode = spec.get("summary_mode")
    selected_window = _normalize_detail_window(selected_window)
    if summary_mode == "status_history_active":
        counts = _status_history_window_counts(filters, predicate=lambda row: row.get("is_active") == "true", as_of=as_of)
    elif summary_mode == "status_history_inactive":
        counts = _status_history_window_counts(filters, predicate=lambda row: row.get("is_inactive") == "true", as_of=as_of)
    elif summary_mode == "status_history_certified":
        counts = _certified_window_counts(filters, as_of)
    elif summary_mode == "current_doctor_login":
        _, rows, _ = _base_rows_for_metric(metric, filters)
        counts = {"Last 24 Hours": 0, "Last Week": 0, "Last Month": 0, "Cumulative": _doctor_login_count(rows)}
    else:
        _, rows, _ = _base_rows_for_metric(metric, filters)
        is_course_metric = clean_text(metric).startswith("doctor_course_") or clean_text(metric).startswith("paramedic_course_")
        counts = _count_window(
            rows,
            date_field=str(spec.get("summary_date_field") or spec.get("date_field") or ""),
            as_of=as_of,
            unique_field=spec.get("summary_unique_field"),
            include_undated_cumulative=is_course_metric,
        )
    return [
        {
            "key": window["key"],
            "label": window["label"],
            "count": counts.get(str(window["label"]), 0),
            "href": _detail_window_href(metric, filters, str(window["key"])),
            "selected": selected_window == window["key"],
        }
        for window in DETAIL_WINDOWS
    ]


def _base_rows_for_metric(metric: str, filters: dict[str, str | None]) -> tuple[dict[str, Any], list[dict[str, Any]], dict[str, Any]]:
    spec = DETAIL_SPECS.get(metric)
    if spec is None:
        raise Http404("Unknown metric")
    rows = filter_rows(_gold_rows(spec["table"], filters), filters)
    if spec.get("predicate"):
        rows = [row for row in rows if spec["predicate"](row)]
    refresh = _latest_refresh(filters) or {}
    return spec, rows, refresh


def _rows_for_metric(metric: str, filters: dict[str, str | None], selected_window: str = "") -> tuple[dict[str, Any], list[dict[str, Any]], dict[str, Any]]:
    spec, rows, refresh = _base_rows_for_metric(metric, filters)
    as_of = parse_date(refresh.get("as_of_date")) or date.today()
    selected_window = _normalize_detail_window(selected_window)
    summary_mode = spec.get("summary_mode")
    if metric in {"patient_videos", "doctor_videos"}:
        if selected_window:
            rows = _filter_rows_by_detail_window(rows, spec.get("summary_date_field"), as_of, selected_window)
        rows = [
            row
            for row in build_video_rankings(_enrich_video_rows(rows))
            if clean_text(row.get("audience")) == ("patient" if metric == "patient_videos" else "doctor")
        ]
    elif metric == "certified_clinics":
        if selected_window:
            rows = _certified_rows_for_window(filters, as_of, selected_window)
        else:
            rows = _derived_certified_rows(
                filters,
                _gold_rows("rpt_doctor_status_current", filters),
                _gold_rows("rpt_course_detail", filters),
            )
    elif summary_mode == "status_history_active" and selected_window:
        rows = _status_history_rows_for_window(filters, predicate=lambda row: row.get("is_active") == "true", as_of=as_of, window_key=selected_window)
    elif summary_mode == "status_history_inactive" and selected_window:
        rows = _status_history_rows_for_window(filters, predicate=lambda row: row.get("is_inactive") == "true", as_of=as_of, window_key=selected_window)
    elif selected_window:
        is_course_metric = clean_text(metric).startswith("doctor_course_") or clean_text(metric).startswith("paramedic_course_")
        rows = _filter_rows_by_detail_window(
            rows,
            spec.get("summary_date_field") or spec.get("date_field"),
            as_of,
            selected_window,
            include_undated_cumulative=is_course_metric,
        )
    elif spec.get("weekly"):
        weekly_start = as_of - timedelta(days=6)
        date_field = spec.get("date_field")
        rows = [
            row
            for row in rows
            if date_field and (parse_date(row.get(date_field)) and weekly_start <= parse_date(row.get(date_field)) <= as_of)
        ]
    date_field = spec.get("date_field")
    if date_field:
        rows.sort(key=lambda row: row.get(date_field) or "", reverse=True)
    elif "rank" in spec["columns"]:
        rows.sort(key=lambda row: _to_int(row.get("rank")), reverse=False)
    return spec, rows, refresh or {}


def detail_context(metric: str, filters: dict[str, str | None], page: int = 1, per_page: int = 25, selected_window: str = "") -> dict[str, Any]:
    selected_window = _normalize_detail_window(selected_window)
    spec, _, refresh = _base_rows_for_metric(metric, filters)
    rows = []
    if selected_window:
        spec, rows, refresh = _rows_for_metric(metric, filters, selected_window=selected_window)
    total_rows = len(rows)
    page = max(page, 1)
    start = (page - 1) * per_page
    end = start + per_page
    page_rows = rows[start:end]
    filters_query = current_filters_query(filters, include_campaign=not clean_text(filters.get("campaign_key")))
    selected_window_query = urlencode({"window": selected_window}) if selected_window else ""
    detail_query_parts = [query for query in [filters_query, selected_window_query] if query]
    detail_query = "&".join(detail_query_parts)
    selected_label = next((str(window["label"]) for window in DETAIL_WINDOWS if window["key"] == selected_window), "")
    return {
        "metric": metric,
        "title": spec["title"],
        "summary_cards": _metric_summary_cards(metric, filters, refresh, selected_window=selected_window),
        "columns": spec["columns"],
        "rows": page_rows,
        "row_count": total_rows,
        "page": page,
        "page_count": max(1, ceil(total_rows / per_page)) if total_rows else 1,
        "filters": filters,
        "filters_query": filters_query,
        "detail_query": detail_query,
        "selected_window": selected_window,
        "selected_window_label": selected_label,
        "has_selected_window": bool(selected_window),
        "route_base": _campaign_route_base(filters),
        "dashboard_href": _campaign_route_base(filters) + (f"?{filters_query}" if filters_query else ""),
        "export_href": f"{_campaign_route_base(filters)}details/{metric}/export/" + (f"?{detail_query}" if detail_query else ""),
        "last_updated": refresh.get("published_at", ""),
        "as_of_date": refresh.get("as_of_date", ""),
    }


def export_detail_csv(metric: str, filters: dict[str, str | None], request: HttpRequest, selected_window: str = "") -> HttpResponse:
    import csv

    spec, rows, _ = _rows_for_metric(metric, filters, selected_window=_normalize_detail_window(selected_window))
    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = f'attachment; filename="sapa-growth-{metric}.csv"'
    writer = csv.writer(response)
    writer.writerow(spec["columns"])
    for row in rows:
        writer.writerow([row.get(column, "") for column in spec["columns"]])
    log_export(metric, f"/sapa-growth/details/{metric}/export/", current_filters_query(filters), len(rows), request.session.session_key)
    return response


def export_certified_csv(global_filters: dict[str, str | None], local_filters: dict[str, str | None], request: HttpRequest) -> HttpResponse:
    import csv

    context = certified_context(global_filters, local_filters)
    if not context["supported"]:
        raise Http404("Certified clinics are not configured")
    rows = context["rows"]
    columns = ["serial_order", "campaign_label", "doctor_display_name", "city", "state", "field_rep_id", "field_rep_name", "certification_status", "certification_date"]
    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = 'attachment; filename="sapa-growth-certified-clinics.csv"'
    writer = csv.writer(response)
    writer.writerow(columns)
    for row in rows:
        writer.writerow([row.get(column, "") for column in columns])
    log_export("certified-clinics", "/sapa-growth/certified-clinics/export/", current_filters_query(local_filters), len(rows), request.session.session_key)
    return response


def export_dashboard_pdf(filters: dict[str, str | None], request: HttpRequest) -> HttpResponse:
    if request.method != "POST":
        return HttpResponseNotAllowed(["POST"])

    snapshot = request.FILES.get("snapshot")
    if snapshot is None:
        return HttpResponseBadRequest("Missing dashboard snapshot")

    image_bytes = snapshot.read()
    if not image_bytes:
        return HttpResponseBadRequest("Empty dashboard snapshot")

    image_reader = ImageReader(BytesIO(image_bytes))
    image_width, image_height = image_reader.getSize()
    if not image_width or not image_height:
        return HttpResponseBadRequest("Invalid dashboard snapshot")

    refresh = _latest_refresh(filters) or {}
    as_of_date = clean_text(refresh.get("as_of_date")) or date.today().isoformat()
    filename = f"sapa-growth-dashboard-{as_of_date}.pdf"

    buffer = BytesIO()
    pdf = pdf_canvas.Canvas(buffer, pagesize=(float(image_width), float(image_height)))
    pdf.drawImage(
        image_reader,
        0,
        0,
        width=float(image_width),
        height=float(image_height),
        preserveAspectRatio=True,
        mask="auto",
    )
    pdf.showPage()
    pdf.save()
    buffer.seek(0)

    response = HttpResponse(buffer.getvalue(), content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    log_export("dashboard-pdf", "/sapa-growth/export/dashboard.pdf", current_filters_query(filters), 0, request.session.session_key)
    return response


def export_dashboard_xlsx(filters: dict[str, str | None], request: HttpRequest) -> HttpResponse:
    context = dashboard_context(filters)
    if not context["ready"]:
        raise Http404("Dashboard data has not been published yet")

    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Summary"
    summary_sheet.append(["Metric", "Value"])
    summary = context["summary"]
    for label, field in (
        ("Webinar Registrations (Weekly)", "webinar_registrations_weekly"),
        ("Onboarded Doctors (Weekly)", "onboarded_doctors_weekly"),
        ("Active Clinics", "active_clinics_current"),
        ("Inactive Clinics", "inactive_clinics_current"),
        ("Total Screenings (Weekly)", "total_screenings_weekly"),
        ("Red Tags (Weekly)", "red_tags_weekly"),
        ("Yellow Tags (Weekly)", "yellow_tags_weekly"),
        ("Follow-ups Scheduled (Weekly)", "followups_scheduled_weekly"),
        ("Reminders Sent (Weekly)", "reminders_sent_weekly"),
    ):
        summary_sheet.append([label, summary.get(field)])

    sheet_specs = [
        ("Active Clinics", _rows_for_metric("active_clinics", filters)[1]),
        ("Inactive Clinics", _rows_for_metric("inactive_clinics", filters)[1]),
        ("Certified Clinics", _rows_for_metric("certified_clinics", filters)[1] if context["certified_supported"] else []),
        ("Total Screenings", _rows_for_metric("total_screenings", filters)[1]),
        ("Red Tags", _rows_for_metric("red_tags", filters)[1]),
        ("Yellow Tags", _rows_for_metric("yellow_tags", filters)[1]),
        ("Followups", _rows_for_metric("followups_scheduled", filters)[1]),
        ("Reminders", _rows_for_metric("reminders_sent", filters)[1]),
        ("Field Rep Logins", _rows_for_metric("field_rep_logins", filters)[1]),
        ("Doctor Course", _rows_for_metric("doctor_course_started", filters)[1] + _rows_for_metric("doctor_course_completed", filters)[1] + _rows_for_metric("doctor_course_pending", filters)[1]),
        ("Paramedic Course", _rows_for_metric("paramedic_course_started", filters)[1] + _rows_for_metric("paramedic_course_completed", filters)[1] + _rows_for_metric("paramedic_course_pending", filters)[1]),
        ("Webinar", _rows_for_metric("webinar_registrations", filters)[1]),
        ("Patient Videos", _rows_for_metric("patient_videos", filters)[1]),
        ("Doctor Videos", _rows_for_metric("doctor_videos", filters)[1]),
    ]

    for title, rows in sheet_specs:
        sheet = workbook.create_sheet(title[:31])
        if rows:
            columns = list(rows[0].keys())
            sheet.append(columns)
            for row in rows:
                sheet.append([row.get(column, "") for column in columns])
        else:
            sheet.append(["No data"])

    filters_sheet = workbook.create_sheet("Filters")
    filters_sheet.append(["Filter", "Value"])
    for key, value in filters.items():
        filters_sheet.append([key, value or ""])
    filters_sheet.append(["As Of Date", context["summary"]["as_of_date"]])
    filters_sheet.append(["Published At", context["summary"]["published_at"]])

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="sapa-growth-dashboard.xlsx"'
    log_export("dashboard-xlsx", "/sapa-growth/export/dashboard.xlsx", current_filters_query(filters), 0, request.session.session_key)
    return response
